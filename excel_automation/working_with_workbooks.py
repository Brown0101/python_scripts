import openpyxl
from openpyxl.styles import *
import os

# Setting up to use a relative path instead of a full path.
dirpath = os.path.dirname(__file__)  # Get the directory path.
filename = os.path.join(dirpath, "Employees.xlsx")  # Add path to filename location

# Lets work with our workbook.
workbook = openpyxl.load_workbook(filename)
print(workbook.properties)  # Returns basic information about the workbook
print(workbook.sheetnames)  # Display all workbook sheet names

# Activate workbook.
workbook.active
sheet = workbook["EmployeeData"]  # Sets the variable to the workbook EmployeeData
workbook.create_sheet("TestSheet")  # Creates a new worksheet
workbook.save(filename)  # Saves the workbook or you can save a new workbook
workbook.remove(sheet)  # Deletes the sheet.

print(sheet.title) # Returns the worksheet's title
print(sheet.active_cell)  # Returns the active cell
print(sheet.dimensions)  # Returns all cell dimensions being used.
print(sheet.sheet_format)  # Get parameters of a sheet
print(sheet.sheet_properties)  # Returns other parameters of the sheet.
print(sheet.sheet_state)  # Returns if the sheet is visible or hidden.
print(sheet.sheet_view)  # Returns more parameters on the sheet.

print(sheet.max_row)  # Returns the number of rows used in the sheet.
print(sheet.max_column)  # Returns the number of columns used in the sheet.

# Return all rows as tuples
for i in sheet.values:
    print(i)

# Two ways to get the value of a cell
sheet['B7'].value
sheet.cell(row=6, column=2).value

# Reference a cell within a sheet
cell = sheet['B9']
cell.row
cell.column
cell.coordinate  # returns B9
cell.data_type  # returns 's' for string
cell.encoding  # In this cass return utf-8 for default
cell.value = 'Frank'
workbook.save(filename)

# Get parent worksheet of cell
cell.parent

# Working with styles
dir(openpyxl.styles)  # Using to review available styles for openpyxl
cell2 = sheet['B8']
font = Font(color=colors.RED, bold=True, italic=True)
fill = PatternFill(fill_type='solid', bgColor='F7FE2E')
border = Border(left = Side(border_style = 'double', color = '322FEC'), right = Side(border_style = 'double',
                color = '322FEC'), top = Side(border_style = 'double', color = '322FEC'),
                bottom = Side(border_style = 'double', color = '322FEC'))
align = Alignment(horizontal='left')
cell.font = font
cell.fill = fill
cell.border = border
cell.alignment = align
workbook.save(filename)